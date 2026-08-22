"""
Pytest versions of the "Today's Trades" Excel export and "Yesterday's
results" feature tests, using the isolated client/db_session fixtures.
"""
from datetime import timedelta

import pytest

from app.models_db import Recommendation, PaperTrade
from app.services.market_hours import today_ist


@pytest.fixture()
def app_client(client, mock_daily_data):
    """The `client` fixture wired up with realistic, internally-consistent mocked market
    data, ready for endpoint-level tests that need the full generation pipeline to run."""
    return client


def test_today_export_xlsx_has_expected_columns_and_rows(app_client):
    r = app_client.get("/api/export-today/xlsx")
    assert r.status_code == 200
    assert r.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert len(r.content) > 500

    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb["Today's Trades"]
    headers = [c.value for c in ws[1]]
    assert headers[:7] == ["Option Name", "Price", "Target 1", "Target 2", "Target 3", "Stop Loss", "Enter When"]
    assert ws.max_row >= 2  # header + at least one trade row


def test_yesterday_with_no_prior_data_is_honest_empty_state(app_client):
    r = app_client.get("/api/tracking/NIFTY/yesterday")
    assert r.status_code == 200
    data = r.json()
    assert data["trade_date"] is None
    assert data["legs"] == []


def test_yesterday_lazily_settles_a_stale_pending_row(app_client, db_session):
    yesterday_date = today_ist() - timedelta(days=1)
    rec = Recommendation(
        index_key="NIFTY", trade_date=yesterday_date, role="PRIMARY", status="PENDING",
        direction="Bullish", option_type="CALL", strike=25800, expiry=(today_ist() + timedelta(days=5)).isoformat(),
        lot_size=75, cmp_at_generation=25750, premium_at_generation=50.0, entry_type="Conditional",
        entry_trigger_desc="Enter on a 15-min close beyond 25800", entry_trigger_index_level=25800,
        target_index_1=25850, target_index_2=25900, target_index_3=25950, stop_index_level=25700,
        target_premium_1=60.0, target_premium_2=70.0, target_premium_3=80.0, stop_premium=30.0,
        atr_pct_at_generation=1.0, confidence_score=65.0, reasoning="test seed", monitor_tick_count=3,
    )
    db_session.add(rec)
    db_session.commit()

    r = app_client.get("/api/tracking/NIFTY/yesterday")
    data = r.json()

    assert data["trade_date"] == yesterday_date.isoformat()
    assert len(data["legs"]) == 1
    leg = data["legs"][0]
    assert leg["status"] == "NOT_EXECUTED", "a stale PENDING from a day that's over should be lazily settled"
    assert "not reached" in leg["not_executed_reason"].lower() or "checked" in leg["not_executed_reason"].lower()


def test_yesterday_shows_accurate_pnl_for_a_closed_profitable_trade(app_client, db_session):
    yesterday_date = today_ist() - timedelta(days=1)
    rec = Recommendation(
        index_key="BANKNIFTY", trade_date=yesterday_date, role="PRIMARY", status="EXECUTED",
        direction="Bullish", option_type="CALL", strike=57000, expiry=(today_ist() + timedelta(days=5)).isoformat(),
        lot_size=30, cmp_at_generation=56900, premium_at_generation=100.0, entry_type="Immediate",
        entry_trigger_desc="Immediate entry", entry_trigger_index_level=56900,
        target_index_1=57100, target_index_2=57200, target_index_3=57300, stop_index_level=56700,
        target_premium_1=130.0, target_premium_2=160.0, target_premium_3=190.0, stop_premium=70.0,
        atr_pct_at_generation=1.0, confidence_score=80.0, reasoning="test seed profitable",
    )
    db_session.add(rec)
    db_session.commit()
    db_session.refresh(rec)

    trade = PaperTrade(
        recommendation_id=rec.id, index_key="BANKNIFTY", trade_date=yesterday_date,
        entry_index_level=56900, entry_premium=100.0, entry_time=today_ist(),
        exit_index_level=57150, exit_premium=135.0, exit_time=today_ist(),
        mfe_premium=135.0, targets_hit=["target_1"], stop_hit=False, status="CLOSED", outcome="TARGET_1",
        pnl_rupees=(135.0 - 100.0) * 30, pnl_pct=round((135.0 - 100.0) / 100.0 * 100, 2),
    )
    db_session.add(trade)
    db_session.commit()

    r = app_client.get("/api/tracking/BANKNIFTY/yesterday")
    leg = r.json()["legs"][0]

    assert leg["status"] == "EXECUTED"
    assert leg["paper_trade"]["status"] == "CLOSED"
    assert leg["paper_trade"]["outcome"] == "TARGET_1"
    assert leg["paper_trade"]["pnl_pct"] == 35.0
    assert leg["paper_trade"]["pnl_rupees"] == pytest.approx(1050.0)


def test_full_endpoint_regression_across_all_indices(app_client):
    """Broad smoke test across every tracking endpoint for every active index."""
    for idx in ["NIFTY", "BANKNIFTY", "SENSEX"]:
        for method, path in [
            ("GET", f"/api/outlook/{idx}"), ("GET", f"/api/tracking/{idx}/today"),
            ("GET", f"/api/tracking/{idx}/yesterday"),
            ("GET", f"/api/tracking/{idx}/paper-trades"), ("GET", f"/api/tracking/{idx}/not-executed"),
            ("GET", f"/api/tracking/{idx}/invalidated"),
            ("GET", f"/api/performance/{idx}"), ("GET", f"/api/performance/{idx}/equity-curve"),
            ("GET", f"/api/export/{idx}/csv"), ("GET", f"/api/export/{idx}/xlsx"),
            ("POST", f"/api/tracking/{idx}/check-now"), ("POST", f"/api/tracking/{idx}/finalize-now"),
        ]:
            r = app_client.get(path) if method == "GET" else app_client.post(path)
            assert r.status_code == 200, f"{method} {path} -> {r.status_code}: {r.text[:300]}"
